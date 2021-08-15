# python
# sendDuesReminders.py - Sends emails based on payment status in spreadsheet

import openpyxl, smtplib, sys

# Open the spreadsheet and get the latest dues status

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
lastCol = sheet.max_solumn
latesMonth = sheet.cell(row=1, column=lastCol).value

# TODO: Check each member's payment status
unpaidMembers = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email
        
# TODO:  Log in to email account
smptObj = smtplib.SMTP('smtp.example.com', 587)
smptObj.ehlo()
smptObj.starttls()
smptObj.login('my_email_adrres@example.com', sys.argv[1])

# TODO: Send out reminder emails
for name, email in unpaidMembers.items():
    body = "Subject = %s dues unpaid. \nDear %s, \nRecords show that you have not paid dues for %s. Please make this payment as soon as possible. Thank you!" % (latesMonth, name, latestMonth)
    print('Sending email to %s...' % email)
    sendmailStatus = smptObj.sendmail('my_email_adress@example.com', email, body)
    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email,sendmailStatus))
smptObj.quit()
